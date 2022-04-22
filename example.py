""" Concurrently parse configuration files and write to Excel workbook """

import glob
import multiprocessing
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import confparser

def convert(value):
    """ Format cell string """
    if isinstance(value, list):
        return ', '.join(map(str, value))
    return str(value) if value else ''

def fill_sheet(ws, data, what, name):
    """ Create named table on sheet from given nested dictionaries """
    # Find unique keys in nested dictionaries
    headers = set(key for vlan in data[what].values() for key in vlan.keys())
    # First row are the headers
    ws.append([what] + list(headers))
    # Iterate over nested dictionaries and append rows
    for key, values in data[what].items():
        # Get items in headers' order and convert to string
        ws.append([key] + [convert(values[key]) for key in headers])
    # Format table
    col = get_column_letter(ws.max_column)
    tab = Table(displayName=name, ref='A1:{}{}'.format(col, ws.max_row))
    style = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def main():
    """ Parse configurations in individual processes """
    dissector = confparser.Dissector.from_file('ios.yaml')
    pool = multiprocessing.Pool()
    result = pool.map(dissector.parse_file, glob.glob('*.cfg'), 1)
    wb = Workbook()
    num = 1
    for data in filter(None, result):
        for what in ('vlan', 'interface'):
            ws = wb.active if num == 1 else wb.create_sheet()
            ws.title = '{} {}'.format(data['hostname'], what)
            fill_sheet(ws, data, what, 'Table{}'.format(num))
            num += 1
    wb.save('output.xlsx')

if __name__ == '__main__':
    main()
